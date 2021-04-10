import openpyxl
import pprint

def conv_sheet_to_list(src, sheet):
    print('opening workbook...')
    wb = openpyxl.load_workbook(src)
    sheet = wb[sheet]
    sheet_data = []
    row_data = []

    print('reading rows...')
    for row in range(2, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            row_data.append(sheet.cell(row, col).value)
        sheet_data.append(row_data)
        row_data = []
    return sheet_data


def conv_sheet_to_py(src, dest, sheet):

    # convert the workbook sheet to a python list object
    sheet_data = conv_sheet_to_list(src, sheet)

    # open a new text file and write the contents of list object to it
    print('writing results...')
    resultFile = open(dest, 'w')
    resultFile.write('data = ' + pprint.pformat(sheet_data))
    resultFile.close()
    print('done...')


if __name__ == "__main__":
    conv_sheet_to_py('t1.xlsx', 'data.py', 'Projects')

    from data import data as d

    print(d[1][1])

